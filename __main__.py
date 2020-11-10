from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, NamedStyle
from datetime import date
from selenium import webdriver
from ECommerWeb.lazada import lazada
from ECommerWeb.shopee import shopee

import os.path
import os

def main():
    product_name = input("Enter the product name (Model num is the best): ")

    browser = webdriver.Chrome()

    # E-commerce
    product_list = lazada().search_product(browser, product_name)
    #product_list.extend(shopee().search_product(browser, product_name))

    #product_list.extend(new_list)
    browser.quit()

    if(len(product_list) > 0):
        create_excel(product_list, product_name)
    else:
        print("No related product found.")


def create_excel(product_list, product_name):
    # Format of excel
    # No. | Title of product | Price | URL link | E-Commerce

    # Define file name
    today = date.today().strftime("%d%m%Y")
    excel_file_name = "ExcelReport/{}_{}.xlsx".format(today, product_name)

    if os.path.exists(excel_file_name):
        workbook = load_workbook(excel_file_name)
        worksheet = workbook.active
    else:
        # Create Excel
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Original Data"

        # Header style
        header_style = NamedStyle(name="hearder_style")
        header_style.font = Font(bold=True)
        header_style.fill = PatternFill(start_color="00FFFF00", end_color="00FFFF00", fill_type="solid")

        # Write header
        header_list = ("No.", "Product Title", "Price", "Url Link", "E-Commerce")
        for col in range(1, len(header_list) + 1):
            header_cell = worksheet.cell(1, col, header_list[col - 1])
            header_cell.style = header_style


    # Write data into worksheet
    product_num = worksheet.max_row
    for data in product_list:
        content = (product_num, data[0], data[1], data[2], data[3])
        worksheet.append(content)
        product_num += 1


    workbook.save(excel_file_name)

    print("Done create excel.")


if __name__ == "__main__":
    main()